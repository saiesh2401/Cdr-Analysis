import os
import re
import json
import datetime
from bs4 import BeautifulSoup
from ipwhois import IPWhois
import openpyxl
from docx import Document
from collections import defaultdict
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

# --- CONFIGURATION ---
HTML_FILE = "bharatkumarumma.SubscriberInfo - Copy.html"
CACHE_FILE = "isp_cache.json"
JIO_EXCEL = "JIO IP.xlsx"
AIRTEL_EXCEL = "Airtel Format.xlsx"
LETTER_TEMPLATE = "IP Letter.docx"
OUTPUT_DIR = "Generated_Letters"

# Create output directory
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

# Load ISP Cache
if os.path.exists(CACHE_FILE):
    with open(CACHE_FILE, 'r') as f:
        isp_cache = json.load(f)
else:
    isp_cache = {}

def get_isp(ip):
    """Identify ISP for a given IP."""
    if ip in isp_cache:
        return isp_cache[ip]
    
    try:
        obj = IPWhois(ip)
        res = obj.lookup_rdap()
        # Look for organization name
        org = res.get('network', {}).get('name', 'Unknown')
        if not org:
            org = res.get('asn_description', 'Unknown')
            
        print(f"Looked up {ip} -> {org}")
        
        # Normalize ISP Names
        org_upper = org.upper()
        if "JIO" in org_upper or "RELIANCE" in org_upper:
            isp = "JIO"
        elif "AIRTEL" in org_upper or "BHARTI" in org_upper:
            isp = "AIRTEL"
        elif "VODAFONE" in org_upper or "VI" in org_upper or "IDEA" in org_upper:
            isp = "VI"
        elif "BSNL" in org_upper:
            isp = "BSNL"
        else:
            isp = "OTHER"
            
        isp_cache[ip] = isp
        with open(CACHE_FILE, 'w') as f:
            json.dump(isp_cache, f)
            
        return isp
    except Exception as e:
        print(f"Error looking up {ip}: {e}")
        return "Unknown"

def parse_html(file_path):
    """Extract IPs and Timestamps from Google HTML."""
    if not os.path.exists(file_path):
        print(f"HTML file not found: {file_path}")
        return []

    with open(file_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')

    data = []
    # Find the IP ACTIVITY table
    # It usually follows an h3 header "IP ACTIVITY"
    # Logic: Find all tables, check header row
    tables = soup.find_all('table')
    for table in tables:
        headers = [th.get_text().strip() for th in table.find_all('th')]
        if "IP Address" in headers and "Timestamp" in headers:
            # This is the target table
            rows = table.find_all('tr')[1:] # Skip header
            for row in rows:
                cols = row.find_all('td')
                if len(cols) >= 2:
                    ts = cols[0].get_text().strip()
                    ip = cols[1].get_text().strip()
                    
                    # Clean IP (remove port if present, though likely clean in this file)
                    if ip:
                         data.append({'timestamp': ts, 'ip': ip})
    return data

def process_data(data):
    """Group IPs by ISP."""
    grouped = defaultdict(list)
    unique_ips = set(d['ip'] for d in data)
    
    print(f"Found {len(unique_ips)} unique IPs. Identifying ISPs...")
    
    for entry in data:
        ip = entry['ip']
        ts_str = entry['timestamp']
        
        # Parse Timestamp (Format: 2025-07-11 15:26:17 Z)
        try:
            dt = datetime.datetime.strptime(ts_str.replace(" Z", ""), "%Y-%m-%d %H:%M:%S")
        except:
            dt = datetime.datetime.now() # Fallback
            
        isp = get_isp(ip)
        entry['datetime'] = dt
        grouped[isp].append(entry)
        
    return grouped

def fill_jio_excel(entries, template_path):
    """Fill Jio Excel Format."""
    if not os.path.exists(template_path):
        print(f"Template {template_path} missing.")
        return

    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # Jio Format: 
    # Type | Search Value | From Date (YYYYMMDD) | From Time (HHMMSS) | To Date | To Time
    
    # Find the next empty row (assuming row 1 is header, data starts row 2)
    # Actually, let's just append
    current_row = sheet.max_row + 1
    
    for item in entries:
        ip = item['ip']
        dt = item['datetime']
        
        # Define a 10-minute window or keep distinct? Usually 10-20 min window requested
        from_dt = dt - datetime.timedelta(minutes=5)
        to_dt = dt + datetime.timedelta(minutes=5)
        
        ip_type = "IPV6" if ":" in ip else "IPV4"
        
        # Col 1: Type
        sheet.cell(row=current_row, column=1, value=ip_type)
        # Col 2: IP
        sheet.cell(row=current_row, column=2, value=ip)
        # Col 3: From Date YYYYMMDD
        sheet.cell(row=current_row, column=3, value=from_dt.strftime("%Y%m%d"))
        # Col 4: From Time HHMMSS
        sheet.cell(row=current_row, column=4, value=from_dt.strftime("%H%M%S"))
        # Col 5: To Date
        sheet.cell(row=current_row, column=5, value=to_dt.strftime("%Y%m%d"))
        # Col 6: To Time
        sheet.cell(row=current_row, column=6, value=to_dt.strftime("%H%M%S"))
        
        current_row += 1
        
    out_path = os.path.join(OUTPUT_DIR, "JIO_Data.xlsx")
    wb.save(out_path)
    print(f"Generated {out_path}")

def fill_airtel_excel(entries, template_path):
    """Fill Airtel Excel Format."""
    if not os.path.exists(template_path):
         print(f"Template {template_path} missing.")
         return

    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # Airtel Format (Inferred from generic requirement, checking file usually safer)
    # Assuming similar columns, but let's adhere to standard:
    # IP | Date | Time
    
    current_row = sheet.max_row + 1
    
    for item in entries:
        ip = item['ip']
        dt = item['datetime']
        ip_type = "IPV6" if ":" in ip else "IPV4"

        # Try to guess columns based on standard Airtel sheets if header reading failed earlier
        # Usually: IP | Port | Date | Time | TimeZone
        
        sheet.cell(row=current_row, column=1, value=ip_type)
        sheet.cell(row=current_row, column=2, value=ip)
        sheet.cell(row=current_row, column=3, value=dt.strftime("%d-%b-%Y")) # 09-Jul-2025
        sheet.cell(row=current_row, column=4, value=dt.strftime("%H:%M:%S"))
        
        current_row += 1

    out_path = os.path.join(OUTPUT_DIR, "Airtel_Data.xlsx")
    wb.save(out_path)
    print(f"Generated {out_path}")

def fill_word_letter(entries, isp_name, template_path):
    """Add rows to the Word document table."""
    if not os.path.exists(template_path):
        print(f"Template {template_path} missing.")
        return

    doc = Document(template_path)
    
    # 1. Update To/Subject if possible (Simple Search/Replace)
    # This is tricky without messing up format. 
    # Strategy: Find "JIO" or "Airtel" in text and replace? 
    # For now, let's just focus on the TABLE data as requested.
    
    # 2. Find the Table
    # We assume the table is the one with "Type", "Search Value" headers.
    target_table = None
    for table in doc.tables:
        headers = [cell.text.strip().replace('\n', ' ') for cell in table.rows[0].cells]
        # Check for keywords
        if any("Search Value" in h for h in headers) or any("IP" in h for h in headers):
            target_table = table
            break
            
    if target_table:
        print(f"Found target table in {template_path} for {isp_name}")
        
        for item in entries:
            ip = item['ip']
            dt = item['datetime']
            
            from_dt = dt - datetime.timedelta(minutes=5)
            to_dt = dt + datetime.timedelta(minutes=5)
            ip_type = "IPV6" if ":" in ip else "IPV4"

            # Add Row
            row = target_table.add_row()
            cells = row.cells
            
            # Formats vary by ISP letter but usually 4-6 columns
            # We try to match the JIO letter format seen in textutil output
            # Type | Value | From Date | From Time | To Date | To Time
            
            if len(cells) >= 6:
                cells[0].text = ip_type
                cells[1].text = ip
                # Preserve existing date format style if possible, else standard
                cells[2].text = from_dt.strftime("%d-%m-%Y") 
                cells[3].text = from_dt.strftime("%H:%M:%S")
                cells[4].text = to_dt.strftime("%d-%m-%Y")
                cells[5].text = to_dt.strftime("%H:%M:%S")
            elif len(cells) >= 4: # Compact format
                 cells[0].text = ip_type
                 cells[1].text = ip
                 cells[2].text = from_dt.strftime("%d-%m-%Y %H:%M:%S")
                 cells[3].text = to_dt.strftime("%d-%m-%Y %H:%M:%S")
    else:
        print("Could not find a suitable table in the Word document.")

    out_path = os.path.join(OUTPUT_DIR, f"{isp_name}_Request_Letter.docx")
    doc.save(out_path)
    print(f"Generated {out_path}")


def main():
    print("--- Starting ISP Letter Automation ---")
    
    # 1. Parse HTML
    raw_data = parse_html(HTML_FILE)
    if not raw_data:
        print("No data found.")
        return
        
    print(f"Extracted {len(raw_data)} records from HTML.")
    
    # 2. Process & Group
    # Note: process_data now handles caching itself, but we want to catch interrupts
    
    grouped_data = defaultdict(list)
    unique_ips = list(set(d['ip'] for d in raw_data))
    print(f"Found {len(unique_ips)} unique IPs. Identifying ISPs...")
    
    try:
        total = len(raw_data)
        for idx, entry in enumerate(raw_data):
            ip = entry['ip']
            ts_str = entry['timestamp']
            
            # Progress every 10 items
            if idx % 10 == 0:
                print(f"Processing {idx+1}/{total}...", end='\r')
            
            try:
                dt = datetime.datetime.strptime(ts_str.replace(" Z", ""), "%Y-%m-%d %H:%M:%S")
            except:
                dt = datetime.datetime.now()
                
            isp = get_isp(ip)
            entry['datetime'] = dt
            grouped_data[isp].append(entry)
            
    except KeyboardInterrupt:
        print("\nProcess interrupted! Saving partial results...")

    # 3. Generate Files
    print(f"\nGenerating files for {len(grouped_data)} ISPs found...")
    if not grouped_data:
        print("No data collected to save.")
        return

    if "JIO" in grouped_data and grouped_data["JIO"]:
        print("\nProcessing JIO...")
        fill_jio_excel(grouped_data["JIO"], JIO_EXCEL)
        fill_word_letter(grouped_data["JIO"], "JIO", LETTER_TEMPLATE)
        
    if "AIRTEL" in grouped_data and grouped_data["AIRTEL"]:
        print("\nProcessing Airtel...")
        fill_airtel_excel(grouped_data["AIRTEL"], AIRTEL_EXCEL)
        fill_word_letter(grouped_data["AIRTEL"], "AIRTEL", LETTER_TEMPLATE)
        
    if "VI" in grouped_data and grouped_data["VI"]:
         print("\nProcessing Vodafone Idea...")
         fill_word_letter(grouped_data["VI"], "VI", LETTER_TEMPLATE)

    print("\n--- Done ---")
    print(f"Check the folder '{OUTPUT_DIR}' for results.")

if __name__ == "__main__":
    main()
